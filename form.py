import os
from tkinter import Tk, Label, Button, Entry, StringVar, messagebox
from openpyxl import Workbook, load_workbook

EXCEL_FILE = 'responses.xlsx'

def intialize_excel():
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Responses"
        sheet.append(["first", "last", "phone"])
        workbook.save(EXCEL_FILE)

def log_response(first, last, email):
    if not os.path.exists(EXCEL_FILE):
        intialize_excel()
    
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([first, last, email])
    workbook.save(EXCEL_FILE)

def submit_form():
    first = first_name.get().strip()
    last = last_name.get().strip()
    email = email_entry.get().strip()

    if not first or not last or not email:
        messagebox.showerror("Error", "All fields are required.")
        return

    log_response(first, last, email)
    messagebox.showinfo("Success", "Response logged successfully!")
    first_name.set("")
    last_name.set("")
    email_entry.set("")
    
intialize_excel()
app = Tk()
app.title("Response Form")
app.geometry("300x200")
first_name = StringVar()
last_name = StringVar()
email_entry = StringVar()
Label(app, text="First Name").pack(pady=5)
first_name_entry = Entry(app, textvariable=first_name)
first_name_entry.pack(pady=5)
Label(app, text="Last Name").pack(pady=5)
last_name_entry = Entry(app, textvariable=last_name)
last_name_entry.pack(pady=5)
Label(app, text="Email").pack(pady=5)
email_entry_field = Entry(app, textvariable=email_entry)
email_entry_field.pack(pady=5)
submit_button = Button(app, text="Submit", command=submit_form)
submit_button.pack(pady=20)
app.mainloop()
